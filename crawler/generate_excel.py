#!/usr/bin/env python3
"""
The Lord of the Skills — Excel Index Generator
================================================
Reads _cache/manifest_final.json, writes a polished .xlsx with:
  - Sheet "Artifacts": every skill file, filterable
  - Sheet "Kingdoms": 10 kingdoms + counts
  - Sheet "Frameworks": framework breakdown
  - Sheet "Canonical": only ⭐ canonical entries
  - Sheet "Stats": aggregate stats
Uses xlsx skill's base.py design tokens.
"""

from __future__ import annotations

import json
import sys
import os
from pathlib import Path
from collections import Counter, defaultdict

# Load xlsx skill base
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
sys.path.insert(0, XLSX_SKILL_DIR)
sys.path.insert(0, os.path.join(XLSX_SKILL_DIR, "templates"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path("/home/z/my-project/the-lord-of-the-skills")
IN_PATH = ROOT / "_cache" / "manifest_final.json"
OUT_PATH = Path("/home/z/my-project/download/the-lord-of-the-skills/Lord_of_the_Skills_Index.xlsx")
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# LOTR theme palette
KINGDOM_COLORS = {
    "gondor":      "1E3A8A",  # royal blue
    "rivendell":   "0F766E",  # teal
    "moria":       "4B5563",  # stone gray
    "lothlorien":  "15803D",  # elven green
    "mordor":      "991B1B",  # dark red
    "the-shire":   "A16207",  # wheat gold
    "isengard":    "52525B",  # iron
    "rohan":       "92400E",  # horse brown
    "fangorn":     "166534",  # bark green
    "mirkwood":    "581C87",  # spider purple
}

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="1E3A8A")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="4B5563")
CANONICAL_FONT = Font(name="Calibri", size=10, bold=True, color="991B1B")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(*[Side(style="thin", color="D1D5DB")] * 4)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_body(ws, start_row=2, end_row=None, end_col=None):
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_artifacts_sheet(wb, files):
    ws = wb.create_sheet("Artifacts")
    # title row
    ws["A1"] = "Lord of the Skills — Complete Artifact Index"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    ws["A2"] = f"{len(files)} artifacts across {len({f['kingdom'] for f in files})} kingdoms and {len({f['framework'] for f in files})} frameworks. ⭐ = canonical."
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:J2")
    # header row
    headers = ["#", "⭐", "Title", "Kingdom", "Domain", "Framework",
               "Skill Type", "Source Repo", "Cluster Size", "Summary"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4)
    ws.row_dimensions[4].height = 24
    # body
    for i, f in enumerate(sorted(files, key=lambda x: (x["kingdom"], x["framework"], x.get("title", "").lower())), 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value="⭐" if f.get("canonical") else "")
        ws.cell(row=r, column=3, value=f.get("title", "(untitled)"))
        ws.cell(row=r, column=4, value=f["kingdom"])
        ws.cell(row=r, column=5, value=f.get("kingdom_label", ""))
        ws.cell(row=r, column=6, value=f["framework"])
        ws.cell(row=r, column=7, value=f.get("skill_type", "other"))
        ws.cell(row=r, column=8, value=f["source_repo"])
        ws.cell(row=r, column=9, value=f.get("cluster_size", 1))
        ws.cell(row=r, column=10, value=f.get("summary", ""))
        # color kingdom cell
        kg_color = KINGDOM_COLORS.get(f["kingdom"], "FFFFFF")
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=kg_color)
        ws.cell(row=r, column=4).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        # canonical marker
        if f.get("canonical"):
            ws.cell(row=r, column=2).font = CANONICAL_FONT
    style_body(ws, start_row=5, end_col=10)
    set_col_widths(ws, [5, 4, 35, 14, 30, 14, 14, 30, 8, 60])
    # freeze
    ws.freeze_panes = "A5"
    # autofilter
    ws.auto_filter.ref = f"A4:J{ws.max_row}"

def write_kingdoms_sheet(wb, files):
    ws = wb.create_sheet("Kingdoms")
    ws["A1"] = "The Ten Kingdoms"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    headers = ["Kingdom", "Domain", "Artifacts", "Canonical", "Motto"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3)
    kingdom_data = {
        "gondor":     ("Gondor", "Coding & Software Engineering", "Gondor sees, Gondor codes."),
        "rivendell":  ("Rivendell", "Research & Knowledge", "Knowledge flows from the Last Homely House."),
        "moria":      ("Moria", "DevOps & Infrastructure", "Speak, friend, and enter the deploy."),
        "lothlorien": ("Lothlórien", "Data & Analysis", "Where data roots run deep."),
        "mordor":     ("Mordor", "Security & Auditing", "One audit to rule them all."),
        "the-shire":  ("The Shire", "Writing & Content", "Quiet writing, deep roots."),
        "isengard":   ("Isengard", "Agents & Orchestration", "Industry wakes the deep."),
        "rohan":      ("Rohan", "Testing & Verification", "Ride now, ride to verify."),
        "fangorn":    ("Fangorn", "Documentation & Memory", "The old forest remembers."),
        "mirkwood":   ("Mirkwood", "Specialized & Niche", "Strange paths through the wood."),
    }
    counts = Counter(f["kingdom"] for f in files)
    canon_counts = Counter(f["kingdom"] for f in files if f.get("canonical"))
    for i, (k, (name, domain, motto)) in enumerate(kingdom_data.items(), 1):
        r = 3 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=domain)
        ws.cell(row=r, column=3, value=counts.get(k, 0))
        ws.cell(row=r, column=4, value=canon_counts.get(k, 0))
        ws.cell(row=r, column=5, value=motto)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=KINGDOM_COLORS[k])
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    style_body(ws, start_row=4, end_col=5)
    set_col_widths(ws, [16, 32, 12, 12, 50])
    ws.freeze_panes = "A4"

def write_frameworks_sheet(wb, files):
    ws = wb.create_sheet("Frameworks")
    ws["A1"] = "Framework Coverage"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    headers = ["Framework", "Artifacts", "Canonical", "Source Repos"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3)
    fw_counts = Counter(f["framework"] for f in files)
    fw_canon = Counter(f["framework"] for f in files if f.get("canonical"))
    fw_repos = defaultdict(set)
    for f in files:
        fw_repos[f["framework"]].add(f["source_repo"])
    for i, (fw, count) in enumerate(fw_counts.most_common(), 1):
        r = 3 + i
        ws.cell(row=r, column=1, value=fw)
        ws.cell(row=r, column=2, value=count)
        ws.cell(row=r, column=3, value=fw_canon.get(fw, 0))
        ws.cell(row=r, column=4, value=len(fw_repos[fw]))
    style_body(ws, start_row=4, end_col=4)
    set_col_widths(ws, [20, 12, 12, 14])
    ws.freeze_panes = "A4"

def write_canonical_sheet(wb, files):
    ws = wb.create_sheet("Canonical ⭐")
    ws["A1"] = "⭐ Canonical Artifacts Only"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = "One representative per concept cluster — the 'greatest hits' of the Lord of the Skills."
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:H2")
    headers = ["#", "Title", "Kingdom", "Framework", "Skill Type",
               "Source Repo", "Cluster Size", "Summary"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4)
    canon = sorted([f for f in files if f.get("canonical")],
                   key=lambda x: (x["kingdom"], x["framework"], x.get("title", "").lower()))
    for i, f in enumerate(canon, 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=f.get("title", "(untitled)"))
        ws.cell(row=r, column=3, value=f["kingdom"])
        ws.cell(row=r, column=4, value=f["framework"])
        ws.cell(row=r, column=5, value=f.get("skill_type", "other"))
        ws.cell(row=r, column=6, value=f["source_repo"])
        ws.cell(row=r, column=7, value=f.get("cluster_size", 1))
        ws.cell(row=r, column=8, value=f.get("summary", ""))
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=KINGDOM_COLORS.get(f["kingdom"], "FFFFFF"))
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    style_body(ws, start_row=5, end_col=8)
    set_col_widths(ws, [5, 35, 14, 14, 14, 30, 10, 60])
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{ws.max_row}"

def write_stats_sheet(wb, manifest, files):
    ws = wb.create_sheet("Stats")
    ws["A1"] = "Lord of the Skills — Compilation Stats"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    rows = [
        ("Generated at",         manifest.get("generated_at", "—")),
        ("Classified at",        manifest.get("classified_at", "—")),
        ("Deduped at",           manifest.get("dedup_at", "—")),
        ("", ""),
        ("Discovered repos",     manifest.get("total_discovered_repos", 0)),
        ("Cloned repos",         manifest.get("total_cloned_repos", 0)),
        ("Total artifacts",      manifest.get("total_extracted_files", 0)),
        ("Canonical (⭐)",       sum(1 for f in files if f.get("canonical"))),
        ("Total clusters",       manifest.get("dedup_stats", {}).get("total_clusters", 0)),
        ("Multi-file clusters",  manifest.get("dedup_stats", {}).get("multi_clusters", 0)),
        ("", ""),
        ("Kingdoms",             len({f["kingdom"] for f in files})),
        ("Frameworks",           len({f["framework"] for f in files})),
        ("Skill types",          len({f.get("skill_type", "other") for f in files})),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    set_col_widths(ws, [25, 50])

def write_about_sheet(wb):
    ws = wb.create_sheet("About", 0)  # make it first
    ws["A1"] = "⚔ The Lord of the Skills ⚔"
    ws["A1"].font = Font(name="Calibri", size=28, bold=True, color="1E3A8A")
    ws.merge_cells("A1:B1")
    ws["A2"] = "One catalog to rule them all, one catalog to find them,"
    ws["A3"] = "One catalog to bring them all, and in the darkness bind them."
    for r in (2, 3):
        ws[f"A{r}"].font = Font(name="Calibri", size=12, italic=True, color="4B5563")
        ws.merge_cells(f"A{r}:B{r}")
    ws["A5"] = "What is this?"
    ws["A5"].font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    ws["A6"] = ("A reusable Python crawler spidered GitHub for agentic AI skill, agent, and rule files "
                "(SKILL.md, AGENTS.md, .cursorrules, .clinerules, .roo/, CONVENTIONS.md, etc.), extracted them, "
                "classified each into one of 10 LOTR-themed kingdoms, tagged by framework, deduplicated with "
                "canonical marking, and packaged them here.")
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A6:B6")
    ws.row_dimensions[6].height = 80
    ws["A8"] = "Sheets in this workbook"
    ws["A8"].font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    sheets_info = [
        ("Artifacts",      "Complete filterable list of every extracted skill file"),
        ("Kingdoms",       "The 10 LOTR kingdoms with artifact counts"),
        ("Frameworks",     "Coverage by source framework (Claude Code, Cursor, Cline, etc.)"),
        ("Canonical ⭐",   "Just the canonical representatives — one per concept"),
        ("Stats",          "Aggregate compilation statistics"),
    ]
    for i, (name, desc) in enumerate(sheets_info, 9):
        ws.cell(row=i, column=1, value=name).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

def main():
    if not IN_PATH.exists():
        print(f"ERROR: {IN_PATH} not found.")
        return
    manifest = json.loads(IN_PATH.read_text())
    files = manifest.get("files", [])
    print(f"Generating Excel index from {len(files)} files...")

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    write_about_sheet(wb)
    write_artifacts_sheet(wb, files)
    write_kingdoms_sheet(wb, files)
    write_frameworks_sheet(wb, files)
    write_canonical_sheet(wb, files)
    write_stats_sheet(wb, manifest, files)

    wb.properties.creator = "Z.ai"
    wb.properties.title = "Lord of the Skills — Index"
    wb.properties.subject = "Agentic AI skill compilation"

    wb.save(OUT_PATH)
    print(f"\n✓ Excel index written: {OUT_PATH}")
    print(f"  Size: {OUT_PATH.stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    main()
